import openpyxl
base = '/Users/meuba/Code/school-research/data/raw/psu_prc/'
for fn in ['multnomah_forecast_2024.xlsx','washington_forecast_2024.xlsx','clackamas_forecast_2024.xlsx','age_sex_county_forecasts.xlsx','ugb_forecasts_all_areas.xlsx']:
    p = base + fn
    print('\n===', fn, '===')
    wb = openpyxl.load_workbook(p, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f'  sheet: {sn!r}  dims={ws.dimensions}  max_row={ws.max_row}  max_col={ws.max_column}')
