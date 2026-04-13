"""Build psu_prc_forecasts.csv from downloaded raw PSU PRC xlsx files.

Output columns:
  area_id, area_name, area_type, county, year, age_group, sex, total_population, source_file, source_note

area_type in {county, ugb_sub_metro, outside_ugb}
age_group: 'total' for totals; otherwise 5-year bands like '0-4', '5-9', ... '85+'.
sex: 'total' (age-sex not broken out here except where derived).

Scope:
- Clackamas, Washington counties: county total + age-banded (PRC age_sex_county_forecasts.xlsx + per-county xlsx)
- Multnomah county: NOT covered by PSU PRC (Metro Council handles it per OPFP). We include UGB small areas only ('Outside UGB Areas' in Multnomah).
- All three counties: Sub-Metro UGB small cities + 'Outside UGB Area'.
"""
import csv
import os
from pathlib import Path
import openpyxl

BASE = Path('/Users/meuba/Code/school-research/data/raw/psu_prc')
OUT = Path('/Users/meuba/Code/school-research/data/raw/psu_prc_forecasts.csv')

rows = []  # list of dicts

# ---- 1. County age-banded forecasts from age_sex master (Clackamas, Washington only) ----
# Master: age_sex_county_forecasts.xlsx -> age_specific_combined_results
# cols: YEAR, COUNTY, SEX, AGE5, POPULATION, SOURCE
p = BASE / 'age_sex_county_forecasts.xlsx'
wb = openpyxl.load_workbook(p, data_only=True)
ws = wb['age_specific_combined_results']
header = None
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        header = row
        continue
    year, county, sex, age5, pop, src = row
    if county not in ('Clackamas', 'Washington'):
        continue
    if year is None or pop is None:
        continue
    rows.append({
        'area_id': f'{county} County',
        'area_name': f'{county} County',
        'area_type': 'county',
        'county': county,
        'year': int(year),
        'age_group': age5,
        'sex': (sex or 'total').lower() if sex else 'total',
        'total_population': round(float(pop), 2),
        'source_file': 'age_sex_county_forecasts.xlsx',
        'source_note': src or '',
    })

# ---- 2. County total population (Clackamas, Washington) from per-county xlsx ----
for cname, fn in [('Clackamas','clackamas_forecast_2024.xlsx'),('Washington','washington_forecast_2024.xlsx')]:
    p = BASE / fn
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb['County Total Population']
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        year, pop = r[0], r[1]
        if year is None or pop is None:
            continue
        rows.append({
            'area_id': f'{cname} County',
            'area_name': f'{cname} County',
            'area_type': 'county',
            'county': cname,
            'year': int(year),
            'age_group': 'total',
            'sex': 'total',
            'total_population': round(float(pop), 2),
            'source_file': fn,
            'source_note': 'PSU PRC Coordinated Forecast (Region 3, 2024)',
        })

# ---- 3. UGB sub-metro small-city + "Outside UGB" forecasts for all 3 counties ----
# From each county file's "UGB Population (all years)" sheet.
for cname, fn in [
    ('Multnomah','multnomah_forecast_2024.xlsx'),
    ('Clackamas','clackamas_forecast_2024.xlsx'),
    ('Washington','washington_forecast_2024.xlsx'),
]:
    p = BASE / fn
    wb = openpyxl.load_workbook(p, data_only=True)
    if 'UGB Population (all years)' not in wb.sheetnames:
        continue
    ws = wb['UGB Population (all years)']
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        ugb, year, pop = r[0], r[1], r[2]
        if ugb is None or year is None or pop is None:
            continue
        area_type = 'outside_ugb' if 'Outside' in str(ugb) else 'ugb_sub_metro'
        rows.append({
            'area_id': f'{cname}|{ugb}',
            'area_name': str(ugb),
            'area_type': area_type,
            'county': cname,
            'year': int(year),
            'age_group': 'total',
            'sex': 'total',
            'total_population': round(float(pop), 2),
            'source_file': fn,
            'source_note': 'PSU PRC UGB forecast (Region 3, 2024). Excludes Portland Metro UGB (see Oregon Metro).',
        })

# ---- 4. UGB master file (cross-county): adds no new sub-Metro UGBs for our 3 counties
#          beyond what's already above, but re-includes quinquennial county totals for Clack/Wash.
#          Skip to avoid duplicates. (Multnomah County is absent from master, as confirmed.)

# ---- Write CSV ----
fields = ['area_id','area_name','area_type','county','year','age_group','sex','total_population','source_file','source_note']
# Sort
rows.sort(key=lambda r: (r['county'], r['area_type'], r['area_name'], r['year'], r['age_group'], r['sex']))
with OUT.open('w', newline='') as f:
    w = csv.DictWriter(f, fieldnames=fields)
    w.writeheader()
    w.writerows(rows)

print(f'wrote {len(rows)} rows to {OUT}')

# Brief summary
from collections import Counter
ctr_county = Counter(r['county'] for r in rows)
ctr_areatype = Counter((r['county'], r['area_type']) for r in rows)
ctr_ages = Counter(r['age_group'] for r in rows)
years = sorted({r['year'] for r in rows})
print('\nBy county:', dict(ctr_county))
print('By (county, area_type):')
for k, v in sorted(ctr_areatype.items()):
    print(f'  {k}: {v}')
print('\nAge groups:', dict(ctr_ages))
print('\nYear range:', min(years), '-', max(years), f'({len(years)} unique years)')
