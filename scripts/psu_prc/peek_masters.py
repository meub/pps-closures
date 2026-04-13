import openpyxl

# Age-sex county forecasts (all counties)
print('=== age_sex_county_forecasts.xlsx  age_specific_combined_results ===')
p = '/Users/meuba/Code/school-research/data/raw/psu_prc/age_sex_county_forecasts.xlsx'
wb = openpyxl.load_workbook(p, data_only=True)
ws = wb['age_specific_combined_results']
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i > 30:
        break
    print(row)

# Count unique counties, years and ages
print('\n--- unique values (sampling) ---')
cols = None
vals_per_col = [set() for _ in range(6)]
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        cols = row
        print('headers:', cols)
        continue
    for j, v in enumerate(row[:6]):
        if len(vals_per_col[j]) < 50:
            vals_per_col[j].add(v)
for j, s in enumerate(vals_per_col):
    print(f'col {j} ({cols[j]}) sample:', list(s)[:30])

# UGB master
print('\n=== ugb_forecasts_all_areas.xlsx  updated_combined_2024 ===')
p2 = '/Users/meuba/Code/school-research/data/raw/psu_prc/ugb_forecasts_all_areas.xlsx'
wb2 = openpyxl.load_workbook(p2, data_only=True)
ws2 = wb2['updated_combined_2024']
for i, row in enumerate(ws2.iter_rows(values_only=True)):
    if i > 20:
        break
    print(row)
