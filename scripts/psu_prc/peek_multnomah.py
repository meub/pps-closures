import openpyxl
p = '/Users/meuba/Code/school-research/data/raw/psu_prc/multnomah_forecast_2024.xlsx'
wb = openpyxl.load_workbook(p, data_only=True)
for sn in wb.sheetnames:
    ws = wb[sn]
    print('===', sn, '===')
    for row in ws.iter_rows(values_only=True):
        print(row)
