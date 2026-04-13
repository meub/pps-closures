import openpyxl
for fn in ['clackamas_forecast_2024.xlsx','washington_forecast_2024.xlsx']:
    p = '/Users/meuba/Code/school-research/data/raw/psu_prc/'+fn
    print('\n=====', fn, '=====')
    wb = openpyxl.load_workbook(p, data_only=True)
    for sn in ['County Total Population', 'County Population By Age', 'UGB Population (select years)', 'UGB Population (all years)']:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        print('\n---', sn, f'({ws.max_row}x{ws.max_column}) ---')
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 15:
                print('  ...')
                break
            print(' ', row)
