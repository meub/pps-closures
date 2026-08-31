"""Recover PPS's functional-capacity denominator from the 2025 School
Utilization Model workbook.

Source: data/raw/PPS SCHOOL UTILIZATION MODEL.xlsx (obtained via a Sunrise
Elementary parent's public-records request, May 2025). The workbook's
`DISTRICT UTILIZATION` tab computes each school's Functional Capacity as
    (Gross Capacity - Student Stations Reduced) x Final Utilization Rate
but the inputs that drive the reductions (SPED focus rooms, Early Childhood
rooms, Title I / TSI / CSI flags) are pulled from an external linked file
that PPS did not include, so those cells and the Functional Capacity column
itself now read #REF!.

We can still recover the exact denominator PPS used: Functional Capacity is a
single fixed number per school, and every historical year divides its
enrollment by that same number. So
    Functional Capacity = enrollment(year) / utilization(year)
The workbook cached those enrollment/utilization pairs from when the model was
intact (2019-20 through 2021-22). The back-out is identical across all three
years for almost every school, confirming an exact recovery. A handful of
schools stored utilization rounded to two decimals, so their recovered value
carries ~1-2% noise and is flagged `fc_approx`.

This is roughly a 2021-vintage figure (the same era as the LRFP-2021 numbers
already on the site). It is NOT a current 2026 recompute: the workbook's live
gross-capacity edits (Benson re-measure, portable removals) don't reach
functional capacity while the reduction inputs are #REF!.

Beyond the recovered capacity, this parser also mines three other tabs of the
same workbook, all joined by Facility ID:
  - CLASSROOM LIST (room-level): 2026-current gross capacity, classroom count,
    average classroom square footage, and counts of science / art-music /
    computer / gym rooms.
  - LEASES: classrooms leased to outside tenants, and the tenant list.
  - K-2 Enrolled: fire-sprinkler status and K-2 phased-sprinkler-exception
    program enrollment.

Output: data/raw/pps_utilization_model_recovered.json, keyed by the ODE
"School Name" as it appears in data/pps_schools.csv, for the merge in
build_master.py.
"""
import json
import re
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string as ci

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "data/raw/PPS SCHOOL UTILIZATION MODEL.xlsx"
OUT = ROOT / "data/raw/pps_utilization_model_recovered.json"

# (enrollment col, utilization col) pairs for the intact, model-computed years.
YEAR_PAIRS = [("AD", "AF"), ("AG", "AH"), ("AI", "AJ")]

# Verified mapping from the workbook's DISTRICT UTILIZATION site name to the
# ODE "School Name" in the master CSV. Built by normalized-name join and
# hand-checked; kept explicit so the provenance is auditable. Workbook rows
# with no in-scope master match (high-school-only sites, closed/admin/special
# buildings, renamed programs) are intentionally omitted.
NAME_MAP = {
    "ABERNETHY": "Abernethy Elementary School",
    "AINSWORTH": "Ainsworth Elementary School",
    "ALAMEDA": "Alameda Elementary School",
    "ARLETA": "Arleta Elementary School",
    "ASTOR": "Astor Elementary School",
    "ATKINSON": "Atkinson Elementary School",
    "BEACH": "Beach Elementary School",
    "BEAUMONT": "Beaumont Middle School",
    "BRIDGER CREATIVE SCIENCE": "Bridger Creative Science School",
    "BRIDLEMILE": "Bridlemile Elementary School",
    "BUCKMAN": "Buckman Elementary School",
    "CAPITOL HILL": "Capitol Hill Elementary School",
    "CHAPMAN": "Chapman Elementary School",
    "CHIEF JOSEPH": "Chief Joseph Elementary School",
    "CLARK": "Clark Elementary School",
    "CLEVELAND": "Cleveland High School",
    "CRESTON": "Creston Elementary School",
    "DAVINCI": "da Vinci Middle School",
    "DUNIWAY": "Duniway Elementary School",
    "FAUBION": "Faubion Elementary School",
    "FOREST PARK": "Forest Park Elementary School",
    "FRANKLIN": "Franklin High School",
    "GEORGE": "George Middle School",
    "GLENCOE": "Glencoe Elementary School",
    "GRANT": "Grant High School",
    "GRAY": "Gray Middle School",
    "GROUT": "Grout Elementary School",
    "HARRISON PARK": "Harrison Park School",
    "HAYHURST": "Hayhurst Elementary School",
    "HOSFORD": "Hosford Middle School",
    "IRVINGTON": "Irvington Elementary School",
    "JACKSON": "Jackson Middle School",
    "JAMES JOHN": "James John Elementary School",
    "JEFFERSON": "Jefferson High School",
    "KELLOGG": "Kellogg Middle School",
    "KELLY": "Kelly Elementary School",
    "LANE": "Lane Middle School",
    "LAURELHURST": "Laurelhurst Elementary School",
    "LEE": "Lee Elementary School",
    "LENT": "Lent Elementary School",
    "LEWIS": "Lewis Elementary School",
    "LINCOLN": "Lincoln High School",
    "LLEWELLYN": "Llewellyn Elementary School",
    "MAPLEWOOD": "Maplewood Elementary School",
    "MARKHAM": "Markham Elementary School",
    "MARYSVILLE": "Marysville Elementary School",
    "MT. TABOR": "Mt Tabor Middle School",
    "OCKLEY GREEN": "Ockley Green Middle School",
    "PENINSULA": "Peninsula Elementary School",
    "RICHMOND": "Richmond Elementary School",
    "RIEKE": "Rieke Elementary School",
    "RIGLER": "Rigler Elementary School",
    "ROOSEVELT": "Roosevelt High School",
    "ROSA PARKS": "Rosa Parks Elementary School",
    "ROSE CITY PARK": "Rose City Park",
    "ROSEWAY HEIGHTS": "Roseway Heights School",
    "SABIN": "Sabin Elementary School",
    "SCOTT": "Scott Elementary School",
    "SELLWOOD": "Sellwood Middle School",
    "SITTON": "Sitton Elementary School",
    "SKYLINE": "Skyline Elementary School",
    "STEPHENSON": "Stephenson Elementary School",
    "VERNON": "Vernon Elementary School",
    "VESTAL": "Vestal Elementary School",
    "WEST SYLVAN": "West Sylvan Middle School",
    "WHITMAN": "Whitman Elementary School",
    "WINTERHAVEN": "Winterhaven School",
    "WOODLAWN": "Woodlawn Elementary School",
    "WOODMERE": "Woodmere Elementary School",
    "WOODSTOCK": "Woodstock Elementary School",
}

# Documented reasons for a capacity discrepancy, from the workbook's own notes.
# Keyed by ODE School Name. Anything not listed gets the generic note in
# build_master.py.
CONFLICT_NOTES = {
    "Creston Elementary School": (
        "The utilization model's classroom count excludes the Creston annex "
        "building; the 2021 LRFP counted those rooms, so its capacity is "
        "roughly double."
    ),
}


def n_decimals(x):
    s = repr(float(x))
    return len(s.split(".")[-1]) if "." in s else 0


def _num(x):
    return x if isinstance(x, (int, float)) else None


def _fid(x):
    """Normalize a Facility ID to a string key (they are ints across tabs)."""
    if x is None:
        return None
    if isinstance(x, float):
        x = int(x)
    return str(x).strip()


def classroom_inventory(wb):
    """Aggregate the room-level CLASSROOM LIST to per-building totals.

    Cols: A Facility ID, F Useable Area, G Current Use Space Class,
    H Gross Capacity. Summing H per facility reproduces the model's gross
    capacity (DISTRICT UTILIZATION col L) exactly, and is 2026-current, so it
    reflects building edits (Benson re-measure, portable removals) that never
    reached functional capacity. Space-class codes carry inconsistent casing,
    matched case-insensitively the way Excel's COUNTIFS does.
    """
    ws = wb["CLASSROOM LIST"]
    agg = {}
    for r in range(2, ws.max_row + 1):
        fid = _fid(ws.cell(r, 1).value)
        if not fid:
            continue
        a = agg.setdefault(fid, {"rooms": 0, "gross": 0, "reg_area": [],
                                 "science": 0, "art_music": 0, "computer": 0, "gym": 0})
        a["rooms"] += 1
        cap = _num(ws.cell(r, 8).value)
        if cap:
            a["gross"] += cap
        cls = str(ws.cell(r, 7).value or "").upper()
        area = _num(ws.cell(r, 6).value)
        if cls.startswith("CR_REG") or "MODULAR" in cls:
            if area:
                a["reg_area"].append(area)
        if "SCI" in cls:
            a["science"] += 1
        elif cls.startswith("CR_ART") or cls.startswith("CR_PERFM"):
            a["art_music"] += 1
        elif "COMP" in cls:
            a["computer"] += 1
        elif "GYM" in cls or cls.startswith("ATHL"):
            a["gym"] += 1
    out = {}
    for fid, a in agg.items():
        out[fid] = {
            "classrooms_total": a["rooms"],
            "gross_capacity": int(a["gross"]),
            "avg_classroom_sqft": (round(sum(a["reg_area"]) / len(a["reg_area"]))
                                   if a["reg_area"] else None),
            "science_rooms": a["science"],
            "art_music_rooms": a["art_music"],
            "computer_rooms": a["computer"],
            "gym_rooms": a["gym"],
        }
    return out


def lease_summary(wb):
    """Aggregate the LEASES tab per building.

    Cols: A Facility ID, C Classrooms, D Tenant, E Tenant Type. Only rows with
    a classroom count reduce usable space; other rows (cell towers, antennas,
    clinics with no room count) are recorded as tenants but not counted as
    leased classrooms.
    """
    ws = wb["LEASES"]
    agg = {}
    for r in range(2, ws.max_row + 1):
        fid = _fid(ws.cell(r, 1).value)
        if not fid or fid == "None":
            continue
        a = agg.setdefault(fid, {"rooms": 0, "tenants": []})
        rooms = _num(ws.cell(r, 3).value)
        if rooms and rooms > 0:
            a["rooms"] += rooms
        tenant = ws.cell(r, 4).value
        ttype = ws.cell(r, 5).value
        if tenant:
            label = str(tenant).strip()
            if ttype and str(ttype).strip() and str(ttype).strip() != label:
                label += f" ({str(ttype).strip()})"
            a["tenants"].append(label)
    out = {}
    for fid, a in agg.items():
        out[fid] = {
            "leased_classrooms": int(a["rooms"]),
            "lease_tenants": "; ".join(a["tenants"][:6]) or None,
        }
    return out


def sprinkler_status(wb):
    """Read the K-2 Enrolled tab per building.

    Cols: B Facility ID, D Enrolled in K-2 Program?, E K-2 Program Expires?,
    F Fully Sprinklered?, G Notes. The K-2 program grants a phased sprinkler
    exception, so most enrolled buildings are not fully sprinklered.
    """
    ws = wb["K-2 Enrolled"]
    out = {}
    for r in range(2, ws.max_row + 1):
        fid = _fid(ws.cell(r, 2).value)
        if not fid:
            continue
        sprink = str(ws.cell(r, 6).value or "").strip()
        k2 = str(ws.cell(r, 4).value or "").strip()
        expires = _num(ws.cell(r, 5).value)
        notes = ws.cell(r, 7).value
        out[fid] = {
            "fully_sprinklered": True if sprink == "Yes" else (False if sprink == "No" else None),
            "k2_program": True if k2 == "Yes" else (False if k2 == "No" else None),
            "k2_program_expires": int(expires) if expires else None,
            "sprinkler_notes": str(notes).strip() if notes else None,
        }
    return out


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["DISTRICT UTILIZATION"]

    inventory = classroom_inventory(wb)
    leases = lease_summary(wb)
    sprinklers = sprinkler_status(wb)

    def val(r, col):
        return ws.cell(r, ci(col)).value

    out = []
    for r in range(2, ws.max_row + 1):
        wb_name = val(r, "C")
        if not wb_name:
            continue
        school_name = NAME_MAP.get(str(wb_name).strip())
        if not school_name:
            continue

        fcs = []
        approx = False
        for en_col, ut_col in YEAR_PAIRS:
            e, u = val(r, en_col), val(r, ut_col)
            if isinstance(e, (int, float)) and isinstance(u, (int, float)) and u:
                fcs.append(e / u)
                if n_decimals(u) <= 2:
                    approx = True
        if not fcs:
            continue

        fid = _fid(val(r, "B"))
        rec = {
            "school_name": school_name,
            "wb_name": str(wb_name).strip(),
            "site_id": fid,
            "config": val(r, "E"),
            "functional_capacity_model": round(sum(fcs) / len(fcs), 1),
            "fc_approx": approx,
            "n_years": len(fcs),
            "conflict_note": CONFLICT_NOTES.get(school_name),
        }
        rec.update(inventory.get(fid, {}))
        rec.update(leases.get(fid, {}))
        rec.update(sprinklers.get(fid, {}))
        out.append(rec)

    OUT.write_text(json.dumps(out, indent=2))
    print(f"Recovered functional capacity for {len(out)} schools -> {OUT}")
    approx = sum(1 for r in out if r["fc_approx"])
    print(f"  {approx} flagged fc_approx (utilization stored at 2-decimal precision)")
    print(f"  inventory: {sum(1 for r in out if r.get('gross_capacity'))} schools")
    print(f"  leases: {sum(1 for r in out if r.get('leased_classrooms'))} schools with leased rooms")
    print(f"  sprinkler: {sum(1 for r in out if r.get('fully_sprinklered') is False)} not fully sprinklered")


if __name__ == "__main__":
    main()
